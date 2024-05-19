import pandas as pd
import warnings
from datetime import datetime
from openpyxl.styles import Border, Side
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import tempfile
import click
from rich.console import Console

warnings.simplefilter(action='ignore', category=pd.errors.PerformanceWarning)

def set_cell_border(sheet):
    """_summary_
    This function sets cell border for all the cells
    in the provided sheet.
    Args:
        sheet (ws): An openpyxl sheet object.
    """
    grey_color_code = "808080"
    border_style = "medium"
    border = Border(
        left=Side(border_style=border_style, color=grey_color_code),
        right=Side(border_style=border_style, color=grey_color_code),
        top=Side(border_style=border_style, color=grey_color_code),
        bottom=Side(border_style=border_style, color=grey_color_code),
    )
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border


def finalize_export_file(wb, export_file_name):
    """_summary_
    This function finalizes the export file by adding border and
    removing unwanted sheets.
    Args:
        wb (Workbook): The workbook object where export excel sheet exists.
        export_file_name (str): The name with which the excel sheet will be saved.
    """

    sheet = wb["Sheet1"]
    # setting border of cells
    set_cell_border(sheet)
    # setting dimension of cells
    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            sheet, min=col, max=col, width=30
        )
    sheet.column_dimensions = dim_holder

    wb.save(export_file_name)


@click.command("Command to fix sheet formatting")
@click.argument("sheet_path", type=click.Path(exists=True), required=True)
def sheet_utility_handler(sheet_path):
    
    console = Console()
    console.print("Reading Data sheet..")
    df = pd.read_excel(sheet_path)

    team = df["Date"][0].title()
    df = df.iloc[2:]
    df.reset_index(drop=True, inplace=True)

    match_fields = ["Date", "Match", "Competition", "Duration"]
    match_fields_df = df[match_fields]

    df = df.drop(match_fields+["Team"], axis=1)

    number_of_matches = len(df)

    report = pd.DataFrame()

    row = 0
    with console.status(f"Processing Data Sheet...", spinner="monkey"):
        import time
        time.sleep(2) # just to test logging
        for i in range(0, number_of_matches, 2):
            for field in match_fields_df.columns.values:
                report.loc[row, field] = match_fields_df[field][i]

            sub_cols = []
            sub_cols_index = 0
            for field in df.columns.values:
                field_name = field
                if "/" in field_name:
                    sub_cols_index = 0
                    sub_cols = field_name.split("/")
                    field_name = sub_cols[0].strip().title()
                elif "Unnamed" in field_name:
                    sub_cols_index += 1
                    if sub_cols_index >= len(sub_cols):
                        continue
                    field_name = sub_cols[0].strip().title() + \
                                " " + sub_cols[sub_cols_index].strip().title()
                report.loc[row, f"{team} {field_name}"] = df[field][i]
                report.loc[row, f"Opponent {field_name}"] = df[field][i+1]
            row += 1 

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        temp_file_name = tmp.name

    # Write the DataFrame to the temporary file
    with pd.ExcelWriter(temp_file_name, engine='openpyxl') as writer:
        report.to_excel(writer, sheet_name='Sheet1', index=False)
        # writer.save()

    # Load the workbook and get the sheet object
    workbook = load_workbook(temp_file_name)

    file_name = sheet_path.split("/")[-1].split(".")[0]
    file_name = f"reports/{file_name}_Edited_{datetime.now()}.xlsx"
    console.print("Generating Data Sheet...")
    finalize_export_file(workbook, file_name)
    console.print(f"File Saved: {file_name}")


if __name__ == '__main__':
    sheet_utility_handler()